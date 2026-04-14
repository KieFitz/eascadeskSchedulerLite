"""Parse uploaded Excel files into structured dicts for Timefold."""

from datetime import date, time
from io import BytesIO
from typing import Any

import openpyxl

from app.services.plan_limits import max_shift_date

EMPLOYEE_REQUIRED_COLS = {"Name", "Skills"}
SHIFT_REQUIRED_COLS = {"Date", "Start Time", "End Time", "Min Staff"}
AVAILABILITY_REQUIRED_COLS = {"Employee", "Type", "Day / Date"}
VALID_AVAIL_TYPES = {"Preferred", "Unpreferred", "Unavailable"}
WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# ── Spanish → English normalisation maps ─────────────────────────────────────

# Sheet name aliases (Spanish → canonical English)
_SHEET_ALIASES: dict[str, str] = {
    "empleados":      "Employees",
    "disponibilidad": "Availability",
    "turnos":         "Shifts",
}

# Column header aliases (lowercase Spanish → canonical English header)
_COLUMN_ALIASES: dict[str, str] = {
    # Employees sheet
    "nombre":              "Name",
    "habilidades":         "Skills",
    "horas mín/semana":    "Min Hours/Week",
    "horas min/semana":    "Min Hours/Week",
    "coste por hora":      "Cost Per Hour",
    # Availability sheet
    "empleado":            "Employee",
    "tipo":                "Type",
    "día / fecha":         "Day / Date",
    "dia / fecha":         "Day / Date",
    "hora inicio":         "Start Time",
    "hora fin":            "End Time",
    # Shifts sheet
    "fecha":               "Date",
    "habilidades requeridas": "Required Skills",
    "personal mín":        "Min Staff",
    "personal min":        "Min Staff",
}

# Weekday aliases (lowercase Spanish → English)
_WEEKDAY_ALIASES: dict[str, str] = {
    "lunes":      "Monday",
    "martes":     "Tuesday",
    "miércoles":  "Wednesday",
    "miercoles":  "Wednesday",
    "jueves":     "Thursday",
    "viernes":    "Friday",
    "sábado":     "Saturday",
    "sabado":     "Saturday",
    "domingo":    "Sunday",
}

# Availability type aliases (lowercase Spanish → English)
_AVAIL_TYPE_ALIASES: dict[str, str] = {
    "preferido":      "Preferred",
    "no preferido":   "Unpreferred",
    "no disponible":  "Unavailable",
}

# Skill aliases — maps lowercase Spanish (and common English variants) to a
# canonical English title-cased form.  Applied to BOTH employee skills and
# shift required-skills so that "Cajero" == "Cashier" after normalisation.
_SKILL_ALIASES: dict[str, str] = {
    # Cashier
    "cajero":           "Cashier",
    "cajera":           "Cashier",
    "caja":             "Cashier",
    # Manager
    "gerente":          "Manager",
    "gestor":           "Manager",
    "gestora":          "Manager",
    "jefe":             "Manager",
    "jefa":             "Manager",
    # Supervisor
    "supervisor":       "Supervisor",
    "supervisora":      "Supervisor",
    # Cook / Chef
    "cocinero":         "Cook",
    "cocinera":         "Cook",
    "cocina":           "Cook",
    "chef":             "Chef",
    # Waiter / Waitress
    "camarero":         "Waiter",
    "mesero":           "Waiter",
    "mozo":             "Waiter",
    "camarera":         "Waitress",
    "mesera":           "Waitress",
    # Bartender / Barista
    "bartender":        "Bartender",
    "barman":           "Bartender",
    "barmaid":          "Bartender",
    "barista":          "Barista",
    # Cleaning
    "limpieza":         "Cleaning",
    "limpiador":        "Cleaning",
    "limpiadora":       "Cleaning",
    # Security
    "seguridad":        "Security",
    "vigilante":        "Security",
    # Receptionist
    "recepcionista":    "Receptionist",
    # Driver
    "conductor":        "Driver",
    "conductora":       "Driver",
    "chófer":           "Driver",
    "chofer":           "Driver",
    # Warehouse / Stock
    "almacén":          "Warehouse",
    "almacen":          "Warehouse",
    "almacenero":       "Warehouse",
    "almacenera":       "Warehouse",
    "reponedor":        "Stock",
    "reponedora":       "Stock",
    # Pharmacist
    "farmacéutico":     "Pharmacist",
    "farmaceutico":     "Pharmacist",
    "farmacéutica":     "Pharmacist",
    "farmaceutica":     "Pharmacist",
    # Nurse
    "enfermero":        "Nurse",
    "enfermera":        "Nurse",
    # Electrician / Plumber
    "electricista":     "Electrician",
    "fontanero":        "Plumber",
    "plomero":          "Plumber",
}


def _normalise_skill(raw: str) -> str:
    """Return the canonical English skill name for the given raw string.

    Looks up the lowercase form in _SKILL_ALIASES first.  If not found,
    falls back to title-casing the raw value so that "cashier" == "Cashier".
    This guarantees that the same concept spelled in Spanish or mixed-case
    always maps to the same string on both the employee and shift sides.
    """
    key = raw.strip().lower()
    return _SKILL_ALIASES.get(key, raw.strip().title())


def _canonical_sheet(name: str) -> str:
    """Return the canonical English sheet name, or the original if not aliased."""
    return _SHEET_ALIASES.get(name.lower().strip(), name)


def _normalise_headers(headers: list[str]) -> list[str]:
    """Map Spanish column headers to their English canonical form."""
    out = []
    for h in headers:
        key = (h or "").strip().lower()
        out.append(_COLUMN_ALIASES.get(key, h))
    return out


def _normalise_weekday(val: str) -> str:
    """Return the English weekday name for either English or Spanish input."""
    return _WEEKDAY_ALIASES.get(val.lower().strip(), val)


def _normalise_avail_type(val: str) -> str:
    """Return the English availability type for either English or Spanish input."""
    return _AVAIL_TYPE_ALIASES.get(val.lower().strip(), val)


def _to_time_str(val: Any) -> str:
    """Normalise an openpyxl time/string cell value to 'HH:MM'."""
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, str):
        return val.strip()[:5]
    return str(val)


def _to_date(val: Any) -> date:
    """Normalise an openpyxl date/string cell value to a date object."""
    if isinstance(val, date):
        return val
    from datetime import datetime
    return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()


def _row_idx(headers: list[str]) -> dict[str, int]:
    return {h: i for i, h in enumerate(headers)}


def parse_excel(file_bytes: bytes, plan: str) -> dict:
    """
    Parse an Excel workbook and return:
    {
        "employees": [...],
        "shifts": [...],
    }
    Raises ValueError on validation errors.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    # Build a canonical-name → sheet object map so Spanish sheet names are accepted
    sheet_map: dict[str, Any] = {_canonical_sheet(n): wb[n] for n in wb.sheetnames}

    if "Employees" not in sheet_map:
        raise ValueError("Missing sheet: 'Employees' (or 'Empleados')")
    if "Shifts" not in sheet_map:
        raise ValueError("Missing sheet: 'Shifts' (or 'Turnos')")

    # Parse availability first so we can attach it to employees by name
    availability: dict[str, dict] = {}  # name.lower() → {preferred, unpreferred, unavailable}
    if "Availability" in sheet_map:
        availability = _parse_availability(sheet_map["Availability"])

    employees = _parse_employees(sheet_map["Employees"], availability)
    shifts = _parse_shifts(sheet_map["Shifts"], plan)

    if not employees:
        raise ValueError("No employees found in the Employees sheet")
    if not shifts:
        raise ValueError("No shifts found in the Shifts sheet")

    return {"employees": employees, "shifts": shifts}


def _parse_employees(ws, availability: dict) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    # Skip hint rows (rows after header where first cell is italic hint text or empty)
    if not rows:
        raise ValueError("Employees sheet is empty")

    headers = _normalise_headers([str(h).strip() if h else "" for h in rows[0]])
    missing = EMPLOYEE_REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(f"Employees sheet missing required columns: {missing}")

    idx = _row_idx(headers)
    employees = []
    for i, row in enumerate(rows[1:], start=2):
        name = row[idx["Name"]]
        if not name or str(name).strip().lower().startswith(("full name", "nombre")):
            continue  # skip hint rows and empty rows

        name_str = str(name).strip()
        name_key = name_str.lower()

        skills_raw = row[idx["Skills"]] if "Skills" in idx else ""
        skills = [_normalise_skill(s) for s in str(skills_raw or "").split(",") if s.strip()]

        min_hours = 0
        if "Min Hours/Week" in idx and row[idx["Min Hours/Week"]]:
            try:
                min_hours = int(row[idx["Min Hours/Week"]])
            except (ValueError, TypeError):
                pass

        cost_per_hour = 0.0
        if "Cost Per Hour" in idx and row[idx["Cost Per Hour"]]:
            try:
                cost_per_hour = float(row[idx["Cost Per Hour"]])
            except (ValueError, TypeError):
                pass

        av = availability.get(name_key, {})

        employees.append({
            "id": name_key.replace(" ", "_"),
            "name": name_str,
            "min_hours_week": min_hours,
            "cost_per_hour": cost_per_hour,
            "skills": skills,
            "preferred_spans": av.get("preferred", []),
            "unpreferred_spans": av.get("unpreferred", []),
            "unavailable_spans": av.get("unavailable", []),
        })
    return employees


def _parse_availability(ws) -> dict:
    """
    Returns a dict keyed by lowercase employee name:
    {
        "alice johnson": {
            "preferred": [{"day": "Monday", "start": "08:00", "end": "14:00"}, ...],
            "unpreferred": [...],
            "unavailable": [...],
        }
    }
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    headers = _normalise_headers([str(h).strip() if h else "" for h in rows[0]])
    missing = AVAILABILITY_REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(f"Availability sheet missing required columns: {missing}")

    idx = _row_idx(headers)
    result: dict[str, dict] = {}

    for i, row in enumerate(rows[1:], start=2):
        employee = row[idx["Employee"]]
        av_type = row[idx["Type"]]
        day_or_date = row[idx["Day / Date"]]

        if not employee or not av_type or not day_or_date:
            continue  # skip empty / hint rows

        employee_str = str(employee).strip()
        # Skip the hint row (English and Spanish)
        if employee_str.lower().startswith(("must match", "debe")):
            continue

        av_type_str = _normalise_avail_type(str(av_type).strip())
        if av_type_str not in VALID_AVAIL_TYPES:
            raise ValueError(
                f"Availability row {i}: Type must be one of {VALID_AVAIL_TYPES}, got '{av_type}'"
            )

        day_str = _normalise_weekday(str(day_or_date).strip())
        # Validate day/date value
        if day_str not in WEEKDAY_NAMES and not _is_date_string(day_str):
            raise ValueError(
                f"Availability row {i}: Day/Date '{day_or_date}' must be a weekday name "
                f"(e.g. 'Monday' / 'Lunes') or a date in YYYY-MM-DD format."
            )

        start_str = None
        end_str = None
        if "Start Time" in idx and row[idx["Start Time"]]:
            start_str = _to_time_str(row[idx["Start Time"]])
        if "End Time" in idx and row[idx["End Time"]]:
            end_str = _to_time_str(row[idx["End Time"]])

        span = {"day": day_str, "start": start_str, "end": end_str}

        key = employee_str.lower()
        if key not in result:
            result[key] = {"preferred": [], "unpreferred": [], "unavailable": []}

        type_key = av_type_str.lower()
        if type_key == "preferred":
            result[key]["preferred"].append(span)
        elif type_key == "unpreferred":
            result[key]["unpreferred"].append(span)
        else:
            result[key]["unavailable"].append(span)

    return result


def _is_date_string(s: str) -> bool:
    try:
        from datetime import datetime
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def _parse_shifts(ws, plan: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Shifts sheet is empty")

    headers = _normalise_headers([str(h).strip() if h else "" for h in rows[0]])
    missing = SHIFT_REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(f"Shifts sheet missing required columns: {missing}")

    idx = _row_idx(headers)
    cutoff = max_shift_date(plan)
    shifts = []
    # Accumulates the next available slot index per (date, start_time) key.
    # Must persist across rows so two rows with the same date+time never produce
    # the same slot number (which would cause a Timefold PlanningId collision).
    slot_counters: dict[str, int] = {}

    for i, row in enumerate(rows[1:], start=2):
        raw_date = row[idx["Date"]]
        raw_start = row[idx["Start Time"]]
        raw_end = row[idx["End Time"]]

        if not raw_date or not raw_start or not raw_end:
            continue
        # Skip hint rows (English: "YYYY-MM-DD format", Spanish: "Formato YYYY-MM-DD")
        if str(raw_date).strip().upper().startswith(("YYYY", "FORMATO")):
            continue

        try:
            shift_date = _to_date(raw_date)
        except Exception:
            raise ValueError(f"Row {i}: invalid date '{raw_date}'. Use YYYY-MM-DD format.")

        if shift_date > cutoff:
            plan_label = "Pro" if plan == "paid" else "Free"
            raise ValueError(
                f"Row {i}: date {shift_date} exceeds the {plan_label} plan limit "
                f"({cutoff}). Upgrade to schedule further ahead."
            )

        min_staff = 1
        if row[idx["Min Staff"]]:
            try:
                min_staff = int(row[idx["Min Staff"]])
            except (ValueError, TypeError):
                pass

        required_skills_raw = ""
        if "Required Skills" in idx:
            required_skills_raw = str(row[idx["Required Skills"]] or "")
        required_skills = [_normalise_skill(s) for s in required_skills_raw.split(",") if s.strip()]

        start_str = _to_time_str(raw_start)
        end_str = _to_time_str(raw_end)
        base_id = f"{shift_date}_{start_str}"

        # Start from the next available slot for this date+time key
        first_slot = slot_counters.get(base_id, 0)
        for offset in range(min_staff):
            global_slot = first_slot + offset
            shifts.append({
                "id": f"{base_id}_slot{global_slot}",
                "date": str(shift_date),
                "start_time": start_str,
                "end_time": end_str,
                "required_skills": required_skills,
                "slot_index": global_slot,
            })
        slot_counters[base_id] = first_slot + min_staff

    return shifts
