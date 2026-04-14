"""Generate the input Excel template, optionally in Spanish."""

from datetime import date, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="4D3DB7")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HINT_FILL    = PatternFill("solid", fgColor="F3F0FF")
HINT_FONT    = Font(italic=True, color="5B6077", size=9)
SECTION_FILL = PatternFill("solid", fgColor="E8DDFF")
AVAIL_COLOURS = {
    "Preferred":   PatternFill("solid", fgColor="D1FAE5"),
    "Unpreferred": PatternFill("solid", fgColor="FEF3C7"),
    "Unavailable": PatternFill("solid", fgColor="FEE2E2"),
}
AVAIL_FONTS = {
    "Preferred":   Font(color="065F46", bold=True, size=10),
    "Unpreferred": Font(color="92400E", bold=True, size=10),
    "Unavailable": Font(color="991B1B", bold=True, size=10),
}

# ── Translations ──────────────────────────────────────────────────────────────
_STRINGS = {
    "en": {
        "sheet_employees":   "Employees",
        "sheet_availability": "Availability",
        "sheet_shifts":      "Shifts",
        "emp_headers": ["Name", "Skills", "Min Hours/Week", "Cost Per Hour"],
        "emp_hints": [
            "Full name (required)",
            "Comma-separated skills, e.g.  cashier, first_aid",
            "Minimum hours per week (optional, leave blank for no minimum)",
            "Hourly rate in your currency (optional)",
        ],
        "av_headers": ["Employee", "Type", "Day / Date", "Start Time", "End Time"],
        "av_hints": [
            "Must match a name in the Employees sheet",
            "Preferred \u00b7 Unpreferred \u00b7 Unavailable",
            'Weekday name (e.g. "Monday") OR specific date (DD-MM-YYYY)',
            "HH:MM \u2014 leave blank for all-day rule",
            "HH:MM \u2014 leave blank for all-day rule",
        ],
        "av_dv_formula": '"Preferred,Unpreferred,Unavailable"',
        "av_dv_error":  "Choose: Preferred, Unpreferred, or Unavailable",
        "av_dv_error_title": "Invalid type",
        "sh_headers": ["Date", "Start Time", "End Time", "Required Skills", "Min Staff"],
        "sh_hints": [
            "DD-MM-YYYY format",
            "HH:MM (24-hour)",
            "HH:MM (24-hour)",
            "Comma-separated skills required (leave blank for any employee)",
            "How many staff needed for this shift (default 1)",
        ],
        # Sample employee names stay in English
        "sample_employees": [
            ["Alice Johnson", "cashier, first_aid, manager",          18, 14.50],
            ["Bob Smith",     "cashier, customer_service",            20, 12.00],
            ["Carol White",   "cashier, customer_service, first_aid", 22, 12.00],
            ["David Brown",   "cashier",                              16, 11.50],
            ["Eve Davis",     "cashier, first_aid",                   21, 15.00],
        ],
    },
    "es": {
        "sheet_employees":    "Empleados",
        "sheet_availability": "Disponibilidad",
        "sheet_shifts":       "Turnos",
        "emp_headers": ["Nombre", "Habilidades", "Horas M\u00edn/Semana", "Coste Por Hora"],
        "emp_hints": [
            "Nombre completo (obligatorio)",
            "Habilidades separadas por comas, ej.  cajero, primeros_auxilios",
            "Horas m\u00ednimas por semana (opcional, dejar en blanco si no hay m\u00ednimo)",
            "Tarifa horaria en tu moneda (opcional)",
        ],
        "av_headers": ["Empleado", "Tipo", "D\u00eda / Fecha", "Hora Inicio", "Hora Fin"],
        "av_hints": [
            "Debe coincidir con un nombre en la hoja Empleados",
            "Preferido \u00b7 No Preferido \u00b7 No Disponible",
            'Nombre del d\u00eda (ej. "Lunes") O fecha espec\u00edfica (DD-MM-YYYY)',
            "HH:MM \u2014 dejar en blanco para regla de todo el d\u00eda",
            "HH:MM \u2014 dejar en blanco para regla de todo el d\u00eda",
        ],
        # Note: availability type VALUES stay English so the backend parser works
        "av_dv_formula": '"Preferred,Unpreferred,Unavailable"',
        "av_dv_error":  "Elige: Preferred, Unpreferred o Unavailable",
        "av_dv_error_title": "Tipo no v\u00e1lido",
        "sh_headers": ["Fecha", "Hora Inicio", "Hora Fin", "Habilidades Requeridas", "Personal M\u00edn"],
        "sh_hints": [
            "Formato DD-MM-YYYY",
            "HH:MM (formato 24h)",
            "HH:MM (formato 24h)",
            "Habilidades requeridas separadas por comas (dejar en blanco para cualquier empleado)",
            "Cu\u00e1ntos empleados se necesitan para este turno (por defecto 1)",
        ],
        "sample_employees": [
            ["Alice Johnson", "cajero, primeros_auxilios, responsable", 18, 14.50],
            ["Bob Smith",     "cajero, atencion_cliente",               20, 12.00],
            ["Carol White",   "cajero, atencion_cliente, primeros_auxilios", 22, 12.00],
            ["David Brown",   "cajero",                                 16, 11.50],
            ["Eve Davis",     "cajero, primeros_auxilios",              21, 15.00],
        ],
    },
}


def _style_header(ws, row: int, cols: list):
    for c, _ in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def _hint_row(ws, row: int, hints: list):
    for c, hint in enumerate(hints, 1):
        cell = ws.cell(row=row, column=c)
        cell.value = hint
        cell.fill = HINT_FILL
        cell.font = HINT_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def build_template_bytes(lang: str = "en") -> bytes:
    """Build the input template workbook and return raw xlsx bytes."""
    s = _STRINGS.get(lang, _STRINGS["en"])
    today = date.today()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Employees ────────────────────────────────────────────────────
    ws_emp = wb.active
    ws_emp.title = s["sheet_employees"]

    ws_emp.append(s["emp_headers"])
    _style_header(ws_emp, 1, s["emp_headers"])
    _hint_row(ws_emp, 2, s["emp_hints"])
    ws_emp.freeze_panes = "A3"

    for row in s["sample_employees"]:
        ws_emp.append(row)
        r = ws_emp.max_row
        if r % 2 == 1:
            for c in range(1, len(s["emp_headers"]) + 1):
                ws_emp.cell(row=r, column=c).fill = SECTION_FILL

    for i, w in enumerate([24, 42, 20, 16], 1):
        ws_emp.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Availability ─────────────────────────────────────────────────
    ws_av = wb.create_sheet(s["sheet_availability"])

    ws_av.append(s["av_headers"])
    _style_header(ws_av, 1, s["av_headers"])
    _hint_row(ws_av, 2, s["av_hints"])
    ws_av.freeze_panes = "A3"

    dv_type = DataValidation(
        type="list",
        formula1=s["av_dv_formula"],
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle=s["av_dv_error_title"],
        error=s["av_dv_error"],
    )
    ws_av.add_data_validation(dv_type)
    dv_type.sqref = "B3:B500"

    sample_avail = [
        ["Alice Johnson",  "Preferred",   "Monday",    "08:00", "14:00"],
        ["Alice Johnson",  "Preferred",   "Tuesday",   "08:00", "14:00"],
        ["Alice Johnson",  "Preferred",   "Wednesday", "08:00", "14:00"],
        ["Alice Johnson",  "Unavailable", "Sunday",    "",      ""     ],
        ["Bob Smith",      "Unavailable", today.strftime("%d-%m-%Y"), "", ""],
        ["Bob Smith",      "Unpreferred", "Friday",    "18:00", "23:00"],
        ["Carol White",    "Preferred",   "Saturday",  "",      ""     ],
        ["Carol White",    "Preferred",   "Sunday",    "",      ""     ],
        ["David Brown",    "Unavailable", "Monday",    "06:00", "12:00"],
        ["David Brown",    "Unavailable", "Tuesday",   "06:00", "12:00"],
    ]
    for row in sample_avail:
        ws_av.append(row)
        r = ws_av.max_row
        av_type = row[1]
        fill = AVAIL_COLOURS.get(av_type)
        font = AVAIL_FONTS.get(av_type)
        if fill and font:
            for c in range(1, len(s["av_headers"]) + 1):
                ws_av.cell(row=r, column=c).fill = fill
            type_cell = ws_av.cell(row=r, column=2)
            type_cell.font = font
            type_cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate([24, 16, 28, 14, 14], 1):
        ws_av.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Shifts ───────────────────────────────────────────────────────
    ws_shifts = wb.create_sheet(s["sheet_shifts"])

    ws_shifts.append(s["sh_headers"])
    _style_header(ws_shifts, 1, s["sh_headers"])
    _hint_row(ws_shifts, 2, s["sh_hints"])
    ws_shifts.freeze_panes = "A3"

    for i in range(7):
        d = (today + timedelta(days=i)).strftime("%d-%m-%Y")
        rows = [
            [d, "08:00", "16:00", "cashier, first_aid", 1],
            [d, "08:00", "16:00", "cashier",             1],
            [d, "14:00", "22:00", "cashier",             1],
        ]
        for row in rows:
            ws_shifts.append(row)
            r = ws_shifts.max_row
            if r % 2 == 1:
                for c in range(1, len(s["sh_headers"]) + 1):
                    ws_shifts.cell(row=r, column=c).fill = SECTION_FILL

    for row in ws_shifts.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate([16, 14, 14, 38, 12], 1):
        ws_shifts.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
