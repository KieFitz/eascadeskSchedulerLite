"""Export a solved schedule result to a user-friendly Excel workbook.

Four sheets:
  1. Employee Schedule — one row per employee, one column per date (Gantt-like)
  2. Shift Schedule   — grouped by date, each shift row shows assigned employee
                        + daily totals (hours, shifts, cost)
  3. All Assignments  — flat editable table (designed for manual tweaks)
  4. Employee Stats   — scheduled hours, shifts worked, and total pay per employee
"""

from collections import defaultdict
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand colours ────────────────────────────────────────────────────────────
PURPLE       = "4D3DB7"
TEAL         = "33D3CF"
DARK         = "14162B"
LAVENDER     = "E8DDFF"
LIGHT_TEAL   = "D4F5F4"
AMBER_LIGHT  = "FEF3C7"
RED_LIGHT    = "FEE2E2"
WHITE        = "FFFFFF"
GREY_LIGHT   = "F3F4F6"

# ── Reusable styles ──────────────────────────────────────────────────────────
HEADER_FILL     = PatternFill("solid", fgColor=PURPLE)
HEADER_FONT     = Font(bold=True, color=WHITE, size=11)
DATE_FILL       = PatternFill("solid", fgColor=DARK)
DATE_FONT       = Font(bold=True, color=WHITE, size=10)
EMP_NAME_FONT   = Font(bold=True, size=10)
ASSIGNED_FILL   = PatternFill("solid", fgColor=LIGHT_TEAL)
UNASSIGNED_FILL = PatternFill("solid", fgColor=RED_LIGHT)
ZEBRA_FILL      = PatternFill("solid", fgColor=GREY_LIGHT)
TOTALS_FILL     = PatternFill("solid", fgColor=LAVENDER)
TOTALS_FONT     = Font(bold=True, color=DARK, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Translations ─────────────────────────────────────────────────────────────
_STRINGS = {
    "en": {
        "sheet1": "Employee Schedule",
        "sheet2": "Shift Schedule",
        "sheet3": "All Assignments",
        "sheet4": "Employee Stats",
        "sheet5": "Substitutes",
        # Sheet 5 column headers
        "s5_date":      "Date",
        "s5_start":     "Start",
        "s5_end":       "End",
        "s5_skills":    "Required Skills",
        "s5_assigned":  "Assigned To",
        "s5_sub1":      "Substitute 1",
        "s5_why1":      "Why",
        "s5_sub2":      "Substitute 2",
        "s5_why2":      "Why",
        "s5_sub3":      "Substitute 3",
        "s5_why3":      "Why",
        "s5_note":      "If the assigned employee calls in sick, these are the next best available substitutes (ranked by skill match, availability, and schedule fit).",
        "employee": "Employee",
        "unassigned": "UNASSIGNED",
        # Sheet 2 column headers
        "s2_start": "Start",
        "s2_end": "End",
        "s2_skills": "Skills",
        "s2_employee": "Employee",
        "s2_cost_hr": "Cost/hr",
        "s2_shift_cost": "Shift Cost",
        # Sheet 2 totals row
        "s2_totals_label": "Day Total",
        "s2_shifts_fmt": "{n} shifts",
        "s2_hours_fmt": "{h:.1f} hrs",
        # Sheet 3 column headers
        "s3_date": "Date",
        "s3_start": "Start Time",
        "s3_end": "End Time",
        "s3_skills": "Required Skills",
        "s3_slot": "Slot #",
        "s3_employee": "Assigned Employee",
        "s3_cost_hr": "Cost/hr",
        "s3_shift_cost": "Shift Cost",
        "s3_dv_error": "Please select an employee from the list",
        "s3_dv_error_title": "Invalid Employee",
        "s3_dv_prompt": "Pick an employee or leave as UNASSIGNED",
        "s3_dv_prompt_title": "Employee",
        # Sheet 4 column headers
        "s4_employee": "Employee",
        "s4_shifts": "Shifts Worked",
        "s4_hours": "Scheduled Hours",
        "s4_pay": "Total Pay",
        "s4_grand_total": "Grand Total",
    },
    "es": {
        "sheet1": "Horario de Empleados",
        "sheet2": "Horario de Turnos",
        "sheet3": "Todas las Asignaciones",
        "sheet4": "Estad\u00edsticas de Empleados",
        "sheet5": "Sustitutos",
        "s5_date":      "Fecha",
        "s5_start":     "Inicio",
        "s5_end":       "Fin",
        "s5_skills":    "Habilidades Requeridas",
        "s5_assigned":  "Asignado a",
        "s5_sub1":      "Sustituto 1",
        "s5_why1":      "Por qu\u00e9",
        "s5_sub2":      "Sustituto 2",
        "s5_why2":      "Por qu\u00e9",
        "s5_sub3":      "Sustituto 3",
        "s5_why3":      "Por qu\u00e9",
        "s5_note":      "Si el empleado asignado llama enfermo, estos son los mejores sustitutos disponibles (clasificados por habilidades, disponibilidad y encaje de horario).",
        "employee": "Empleado",
        "unassigned": "SIN ASIGNAR",
        # Sheet 2
        "s2_start": "Inicio",
        "s2_end": "Fin",
        "s2_skills": "Habilidades",
        "s2_employee": "Empleado",
        "s2_cost_hr": "Coste/h",
        "s2_shift_cost": "Coste del Turno",
        "s2_totals_label": "Total del D\u00eda",
        "s2_shifts_fmt": "{n} turnos",
        "s2_hours_fmt": "{h:.1f} h",
        # Sheet 3
        "s3_date": "Fecha",
        "s3_start": "Hora Inicio",
        "s3_end": "Hora Fin",
        "s3_skills": "Habilidades Requeridas",
        "s3_slot": "Puesto",
        "s3_employee": "Empleado Asignado",
        "s3_cost_hr": "Coste/h",
        "s3_shift_cost": "Coste del Turno",
        "s3_dv_error": "Por favor selecciona un empleado de la lista",
        "s3_dv_error_title": "Empleado no v\u00e1lido",
        "s3_dv_prompt": "Elige un empleado o deja como SIN ASIGNAR",
        "s3_dv_prompt_title": "Empleado",
        # Sheet 4
        "s4_employee": "Empleado",
        "s4_shifts": "Turnos Trabajados",
        "s4_hours": "Horas Programadas",
        "s4_pay": "Salario Total",
        "s4_grand_total": "Total General",
    },
}


def _s(lang: str, key: str) -> str:
    return _STRINGS.get(lang, _STRINGS["en"]).get(key, _STRINGS["en"].get(key, key))


def _apply_header(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_date_banner(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = DATE_FILL
        cell.font = DATE_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def _auto_width(ws, col: int, min_w: float = 10, max_w: float = 28):
    letter = get_column_letter(col)
    best = min_w
    for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
        cell = row[0]
        if cell.value:
            best = max(best, min(len(str(cell.value)) + 3, max_w))
    ws.column_dimensions[letter].width = best


def _calc_hours(start: str, end: str) -> float:
    """Calculate shift duration in hours from HH:MM strings."""
    try:
        sh, sm = map(int, start.split(":"))
        eh, em = map(int, end.split(":"))
        minutes = eh * 60 + em - sh * 60 - sm
        if minutes < 0:
            minutes += 24 * 60  # overnight shift
        return minutes / 60
    except (ValueError, AttributeError):
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
#  PUBLIC ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def build_schedule_excel(
    employees: list[dict],
    shifts: list[dict],
    assignments: list[dict],
    lang: str = "en",
) -> bytes:
    wb = openpyxl.Workbook()

    _build_employee_schedule(wb, employees, assignments, lang)
    _build_shift_schedule(wb, assignments, lang)
    _build_all_assignments(wb, employees, assignments, lang)
    _build_employee_stats(wb, employees, assignments, lang)
    _build_substitutes(wb, employees, shifts, assignments, lang)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — Employee Schedule (Gantt-like)
# ─────────────────────────────────────────────────────────────────────────────

def _build_employee_schedule(wb, employees: list[dict], assignments: list[dict], lang: str):
    ws = wb.active
    ws.title = _s(lang, "sheet1")

    unassigned_label = _s(lang, "unassigned")

    emp_shifts: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    all_dates: set[str] = set()

    for a in assignments:
        name = a.get("employee_name") or unassigned_label
        d = a.get("date", "")
        start = a.get("start_time", "")
        end = a.get("end_time", "")
        all_dates.add(d)
        emp_shifts[name][d].append(f"{start}\u2013{end}")

    sorted_dates = sorted(all_dates)
    emp_names = [e.get("name", "") for e in employees]

    ws.cell(row=1, column=1, value=_s(lang, "employee"))
    for j, d in enumerate(sorted_dates, start=2):
        ws.cell(row=1, column=j, value=d)
    _apply_header(ws, 1, 1 + len(sorted_dates))

    for i, name in enumerate(emp_names, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=1).font = EMP_NAME_FONT
        ws.cell(row=i, column=1).alignment = LEFT
        ws.cell(row=i, column=1).border = THIN_BORDER

        for j, d in enumerate(sorted_dates, start=2):
            cell = ws.cell(row=i, column=j)
            shift_list = emp_shifts.get(name, {}).get(d, [])
            if shift_list:
                cell.value = "\n".join(shift_list)
                cell.fill = ASSIGNED_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        if i % 2 == 0:
            for j in range(1, 2 + len(sorted_dates)):
                c = ws.cell(row=i, column=j)
                if not c.fill or c.fill.fgColor.rgb == "00000000":
                    c.fill = ZEBRA_FILL

    unassigned_shifts = emp_shifts.get(unassigned_label, {})
    if unassigned_shifts:
        r = len(emp_names) + 2
        ws.cell(row=r, column=1, value=unassigned_label)
        ws.cell(row=r, column=1).font = Font(bold=True, color="DC2626", size=10)
        ws.cell(row=r, column=1).border = THIN_BORDER
        for j, d in enumerate(sorted_dates, start=2):
            cell = ws.cell(row=r, column=j)
            shift_list = unassigned_shifts.get(d, [])
            if shift_list:
                cell.value = "\n".join(shift_list)
                cell.fill = UNASSIGNED_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    for j in range(2, 2 + len(sorted_dates)):
        ws.column_dimensions[get_column_letter(j)].width = 16

    ws.freeze_panes = "B2"


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — Shift Schedule (grouped by date, with daily totals)
# ─────────────────────────────────────────────────────────────────────────────

def _build_shift_schedule(wb, assignments: list[dict], lang: str):
    ws = wb.create_sheet(_s(lang, "sheet2"))

    unassigned_label = _s(lang, "unassigned")

    by_date: dict[str, list[dict]] = defaultdict(list)
    for a in assignments:
        by_date[a.get("date", "")].append(a)

    sorted_dates = sorted(by_date.keys())

    cols = [
        _s(lang, "s2_start"),
        _s(lang, "s2_end"),
        _s(lang, "s2_skills"),
        _s(lang, "s2_employee"),
        _s(lang, "s2_cost_hr"),
        _s(lang, "s2_shift_cost"),
    ]
    num_cols = len(cols)

    row_idx = 1
    for d in sorted_dates:
        # Date banner
        ws.cell(row=row_idx, column=1, value=d)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
        _apply_date_banner(ws, row_idx, num_cols)
        row_idx += 1

        # Column headers
        for c, h in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=c, value=h)
        _apply_header(ws, row_idx, num_cols)
        row_idx += 1

        day_shifts = sorted(
            by_date[d],
            key=lambda x: (x.get("start_time", ""), x.get("slot_index", 0)),
        )

        day_total_hours = 0.0
        day_total_cost  = 0.0
        day_shift_count = len(day_shifts)

        for a in day_shifts:
            skills = a.get("required_skills", [])
            skills_str = ", ".join(skills) if isinstance(skills, list) else str(skills)
            emp_name = a.get("employee_name") or unassigned_label
            cost_h = a.get("cost_per_hour", 0)
            shift_cost = a.get("shift_cost", 0)

            hours = _calc_hours(a.get("start_time", ""), a.get("end_time", ""))
            day_total_hours += hours
            day_total_cost  += shift_cost or 0

            row_data = [
                a.get("start_time", ""),
                a.get("end_time", ""),
                skills_str,
                emp_name,
                f"\u20ac{cost_h:.2f}" if cost_h else "",
                f"\u20ac{shift_cost:.2f}" if shift_cost else "",
            ]
            for c, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=c, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER

            if emp_name == unassigned_label:
                for c in range(1, num_cols + 1):
                    ws.cell(row=row_idx, column=c).fill = UNASSIGNED_FILL
            elif row_idx % 2 == 0:
                for c in range(1, num_cols + 1):
                    ws.cell(row=row_idx, column=c).fill = ZEBRA_FILL

            row_idx += 1

        # Daily totals row
        shifts_fmt = _s(lang, "s2_shifts_fmt").format(n=day_shift_count)
        hours_fmt  = _s(lang, "s2_hours_fmt").format(h=day_total_hours)
        totals_data = [
            _s(lang, "s2_totals_label"),
            "",
            shifts_fmt,
            hours_fmt,
            "",
            f"\u20ac{day_total_cost:.2f}",
        ]
        for c, val in enumerate(totals_data, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill = TOTALS_FILL
            cell.font = TOTALS_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row_idx += 1

        # Spacer
        row_idx += 1

    widths = [10, 10, 30, 22, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 3 — All Assignments (flat, editable)
# ─────────────────────────────────────────────────────────────────────────────

def _build_all_assignments(wb, employees: list[dict], assignments: list[dict], lang: str):
    ws = wb.create_sheet(_s(lang, "sheet3"))

    unassigned_label = _s(lang, "unassigned")

    headers = [
        _s(lang, "s3_date"),
        _s(lang, "s3_start"),
        _s(lang, "s3_end"),
        _s(lang, "s3_skills"),
        _s(lang, "s3_slot"),
        _s(lang, "s3_employee"),
        _s(lang, "s3_cost_hr"),
        _s(lang, "s3_shift_cost"),
    ]
    ws.append(headers)
    _apply_header(ws, 1, len(headers))

    emp_names = sorted({e.get("name", "") for e in employees})
    emp_list_str = ",".join(emp_names) if emp_names else ""

    sorted_assignments = sorted(
        assignments,
        key=lambda x: (x.get("date", ""), x.get("start_time", ""), x.get("slot_index", 0)),
    )

    for i, a in enumerate(sorted_assignments):
        skills = a.get("required_skills", [])
        skills_str = ", ".join(skills) if isinstance(skills, list) else str(skills)
        emp_name = a.get("employee_name") or unassigned_label
        cost_h = a.get("cost_per_hour", 0)
        shift_cost = a.get("shift_cost", 0)

        row = [
            a.get("date", ""),
            a.get("start_time", ""),
            a.get("end_time", ""),
            skills_str,
            (a.get("slot_index") or 0) + 1,
            emp_name,
            cost_h,
            shift_cost,
        ]
        ws.append(row)
        r = ws.max_row

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        if emp_name == unassigned_label:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNASSIGNED_FILL
        elif i % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = ZEBRA_FILL

    if emp_list_str and len(sorted_assignments) > 0:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1=f'"{emp_list_str}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error       = _s(lang, "s3_dv_error")
        dv.errorTitle  = _s(lang, "s3_dv_error_title")
        dv.prompt      = _s(lang, "s3_dv_prompt")
        dv.promptTitle = _s(lang, "s3_dv_prompt_title")
        last_row = 1 + len(sorted_assignments)
        dv.add(f"F2:F{last_row}")
        ws.add_data_validation(dv)

    widths = [14, 13, 13, 30, 8, 24, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 4 — Employee Stats (scheduled hours and pay)
# ─────────────────────────────────────────────────────────────────────────────

def _build_employee_stats(wb, employees: list[dict], assignments: list[dict], lang: str):
    ws = wb.create_sheet(_s(lang, "sheet4"))

    headers = [
        _s(lang, "s4_employee"),
        _s(lang, "s4_shifts"),
        _s(lang, "s4_hours"),
        _s(lang, "s4_pay"),
    ]
    ws.append(headers)
    _apply_header(ws, 1, len(headers))

    # Aggregate per employee (named employees only, skip unassigned)
    emp_names_ordered = [e.get("name", "") for e in employees]
    stats: dict[str, dict] = {
        name: {"shifts": 0, "hours": 0.0, "pay": 0.0}
        for name in emp_names_ordered
        if name
    }

    for a in assignments:
        name = a.get("employee_name")
        if not name or name not in stats:
            continue
        hours = _calc_hours(a.get("start_time", ""), a.get("end_time", ""))
        stats[name]["shifts"] += 1
        stats[name]["hours"]  += hours
        stats[name]["pay"]    += a.get("shift_cost", 0) or 0

    grand_shifts = 0
    grand_hours  = 0.0
    grand_pay    = 0.0

    for i, name in enumerate(emp_names_ordered):
        if not name or name not in stats:
            continue
        row_stats = stats[name]
        grand_shifts += row_stats["shifts"]
        grand_hours  += row_stats["hours"]
        grand_pay    += row_stats["pay"]

        row_data = [
            name,
            row_stats["shifts"],
            round(row_stats["hours"], 2),
            f"\u20ac{row_stats['pay']:.2f}",
        ]
        ws.append(row_data)
        r = ws.max_row

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if c == 1:
                cell.font = EMP_NAME_FONT
                cell.alignment = LEFT

        if i % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = ZEBRA_FILL

    # Grand totals row
    totals_row = [
        _s(lang, "s4_grand_total"),
        grand_shifts,
        round(grand_hours, 2),
        f"\u20ac{grand_pay:.2f}",
    ]
    ws.append(totals_row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = TOTALS_FILL
        cell.font = TOTALS_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = LEFT

    widths = [26, 18, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 5 — Substitutes (sick-call quick-reference)
# ─────────────────────────────────────────────────────────────────────────────

def _to_mins(t: str) -> int:
    try:
        h, m = t.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 0


def _rank_substitutes_excel(
    employees: list[dict],
    shifts: list[dict],
    assignments: list[dict],
    target_shift: dict,
    max_results: int = 3,
) -> list[dict]:
    """Simple substitute ranking for the Excel sheet.

    Factors: skill match, same-day schedule overlap.
    (Availability spans are included if present in employee dicts.)
    """
    shift_id   = target_shift.get("id", "")
    shift_date = target_shift.get("date", "")
    s_start    = _to_mins(target_shift.get("start_time", ""))
    s_end      = _to_mins(target_shift.get("end_time", "") )
    if s_end <= s_start:
        s_end += 1440
    required = set(target_shift.get("required_skills") or [])

    # Same-day overlap map
    shift_by_id = {s.get("id", ""): s for s in shifts}
    overlapping_emps: set[str] = set()
    for a in assignments:
        eid = a.get("employee_id")
        sid = a.get("shift_id")
        if not eid or sid == shift_id:
            continue
        other = shift_by_id.get(sid)
        if not other or other.get("date") != shift_date:
            continue
        o_s = _to_mins(other.get("start_time", ""))
        o_e = _to_mins(other.get("end_time", ""))
        if o_e <= o_s:
            o_e += 1440
        if s_start < o_e and s_end > o_s:
            overlapping_emps.add(eid)

    results = []
    for emp in employees:
        eid       = emp.get("id", "")
        emp_skills = set(emp.get("skills") or [])
        skills_ok  = not required or required.issubset(emp_skills)
        missing    = sorted(required - emp_skills)
        overlaps   = eid in overlapping_emps

        score = (4 if skills_ok else 0) + (-10 if overlaps else 0)

        reasons = []
        if not skills_ok:
            reasons.append(f"Missing: {', '.join(missing)}")
        if overlaps:
            reasons.append("Already scheduled")
        if not reasons:
            reasons.append("Available")

        results.append({
            "name":   emp.get("name", ""),
            "score":  score,
            "reason": " · ".join(reasons),
        })

    results.sort(key=lambda x: (-x["score"], x["name"]))
    return results[:max_results]


def _build_substitutes(
    wb,
    employees: list[dict],
    shifts: list[dict],
    assignments: list[dict],
    lang: str,
) -> None:
    ws = wb.create_sheet(_s(lang, "sheet5"))

    # Introductory note spanning all columns
    note_row = 1
    num_cols = 11
    ws.cell(row=note_row, column=1, value=_s(lang, "s5_note"))
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=num_cols)
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.font = Font(italic=True, color="6B7280", size=9)
    ws.row_dimensions[note_row].height = 30

    # Header row
    header_row = 2
    headers = [
        _s(lang, "s5_date"),
        _s(lang, "s5_start"),
        _s(lang, "s5_end"),
        _s(lang, "s5_skills"),
        _s(lang, "s5_assigned"),
        _s(lang, "s5_sub1"),
        _s(lang, "s5_why1"),
        _s(lang, "s5_sub2"),
        _s(lang, "s5_why2"),
        _s(lang, "s5_sub3"),
        _s(lang, "s5_why3"),
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    _apply_header(ws, header_row, num_cols)

    # Build an assignment lookup: shift_id → assignment
    assign_by_shift = {a.get("shift_id", ""): a for a in assignments}

    # Sort all shifts by date then start time
    unassigned_label = _s(lang, "unassigned")
    sorted_shifts = sorted(shifts, key=lambda s: (s.get("date", ""), s.get("start_time", "")))

    row_idx = header_row + 1
    for i, shift in enumerate(sorted_shifts):
        a = assign_by_shift.get(shift.get("id", ""), {})
        assigned_name = a.get("employee_name") or unassigned_label
        skills        = shift.get("required_skills") or []
        skills_str    = ", ".join(skills)

        subs = _rank_substitutes_excel(employees, shifts, assignments, shift, max_results=3)

        row_data: list = [
            shift.get("date", ""),
            shift.get("start_time", ""),
            shift.get("end_time", ""),
            skills_str,
            assigned_name,
        ]
        for sub in subs:
            row_data += [sub["name"], sub["reason"]]
        while len(row_data) < num_cols:
            row_data.append("")

        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

        # Highlight unassigned shifts in amber
        if not a.get("employee_id"):
            for c in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor="FEF3C7")
        elif i % 2 == 0:
            for c in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=c).fill = ZEBRA_FILL

        row_idx += 1

    # Column widths
    widths = [14, 10, 10, 28, 22, 22, 30, 22, 30, 22, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
