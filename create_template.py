"""
Run this script once to generate the sample Excel template.
Output: frontend/public/schedule_template.xlsx
"""
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="4D3DB7")   # brand-purple
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HINT_FILL     = PatternFill("solid", fgColor="F3F0FF")   # light lavender
HINT_FONT     = Font(italic=True, color="5B6077", size=9)
SECTION_FILL  = PatternFill("solid", fgColor="E8DDFF")   # lavender alt row
AVAIL_COLOURS = {
    "Preferred":   PatternFill("solid", fgColor="D1FAE5"),  # green-100
    "Unpreferred": PatternFill("solid", fgColor="FEF3C7"),  # amber-100
    "Unavailable": PatternFill("solid", fgColor="FEE2E2"),  # red-100
}
AVAIL_FONTS = {
    "Preferred":   Font(color="065F46", bold=True, size=10),
    "Unpreferred": Font(color="92400E", bold=True, size=10),
    "Unavailable": Font(color="991B1B", bold=True, size=10),
}


def _style_header(ws, row: int, cols: list[str]):
    for c, _ in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22


def _hint_row(ws, row: int, hints: list[str]):
    for c, hint in enumerate(hints, 1):
        cell = ws.cell(row=row, column=c)
        cell.value = hint
        cell.fill = HINT_FILL
        cell.font = HINT_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def create_template():
    wb = openpyxl.Workbook()

    today = date.today()

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Employees
    # ══════════════════════════════════════════════════════════════════════════
    ws_emp = wb.active
    ws_emp.title = "Employees"

    emp_headers = ["Name", "Skills", "Min Hours/Week", "Cost Per Hour"]
    emp_hints = [
        "Full name (required)",
        "Comma-separated skills, e.g.  cashier, first_aid",
        "Minimum hours per week (optional, leave blank for no minimum)",
        "Hourly rate in your currency (optional)",
    ]
    ws_emp.append(emp_headers)
    _style_header(ws_emp, 1, emp_headers)
    _hint_row(ws_emp, 2, emp_hints)
    _freeze(ws_emp, "A3")

    sample_employees = [
        ["Alice Johnson", "cashier, first_aid, manager",          32, 14.50],
        ["Bob Smith",     "cashier, customer_service",            20, 12.00],
        ["Carol White",   "cashier, customer_service, first_aid", 24, 12.00],
        ["David Brown",   "cashier",                              16, 11.50],
        ["Eve Davis",     "cashier, first_aid",                   40, 15.00],
    ]
    for row in sample_employees:
        ws_emp.append(row)
        # zebra stripe
        r = ws_emp.max_row
        if r % 2 == 1:
            for c in range(1, len(emp_headers) + 1):
                ws_emp.cell(row=r, column=c).fill = SECTION_FILL

    col_widths = [24, 42, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws_emp.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Availability
    # ══════════════════════════════════════════════════════════════════════════
    ws_av = wb.create_sheet("Availability")

    av_headers = ["Employee", "Type", "Day / Date", "Start Time", "End Time"]
    av_hints = [
        "Must match a name in the Employees sheet",
        "Preferred · Unpreferred · Unavailable",
        'Weekday name (e.g. "Monday") OR specific date (YYYY-MM-DD)',
        "HH:MM — leave blank for all-day rule",
        "HH:MM — leave blank for all-day rule",
    ]
    ws_av.append(av_headers)
    _style_header(ws_av, 1, av_headers)
    _hint_row(ws_av, 2, av_hints)
    _freeze(ws_av, "A3")

    # Dropdown validation for Type column (B)
    dv_type = DataValidation(
        type="list",
        formula1='"Preferred,Unpreferred,Unavailable"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid type",
        error="Choose: Preferred, Unpreferred, or Unavailable",
    )
    ws_av.add_data_validation(dv_type)
    dv_type.sqref = "B3:B500"

    sample_avail = [
        # Alice prefers mornings, unavailable Sundays
        ["Alice Johnson",  "Preferred",   "Monday",    "08:00", "14:00"],
        ["Alice Johnson",  "Preferred",   "Tuesday",   "08:00", "14:00"],
        ["Alice Johnson",  "Preferred",   "Wednesday", "08:00", "14:00"],
        ["Alice Johnson",  "Unavailable", "Sunday",    "",      ""     ],
        # Bob unavailable on a specific date, unprefers late evenings
        ["Bob Smith",      "Unavailable", today.strftime("%Y-%m-%d"), "", ""],
        ["Bob Smith",      "Unpreferred", "Friday",    "18:00", "23:00"],
        # Carol prefers weekends
        ["Carol White",    "Preferred",   "Saturday",  "",      ""     ],
        ["Carol White",    "Preferred",   "Sunday",    "",      ""     ],
        # David unavailable mornings
        ["David Brown",    "Unavailable", "Monday",    "06:00", "12:00"],
        ["David Brown",    "Unavailable", "Tuesday",   "06:00", "12:00"],
    ]

    for row in sample_avail:
        ws_av.append(row)
        r = ws_av.max_row
        av_type = row[1]
        fill = AVAIL_COLOURS.get(av_type)
        font = AVAIL_FONTS.get(av_type)
        if fill and font:
            for c in range(1, len(av_headers) + 1):
                ws_av.cell(row=r, column=c).fill = fill
            # Bold+coloured type cell
            type_cell = ws_av.cell(row=r, column=2)
            type_cell.font = font
            type_cell.alignment = Alignment(horizontal="center")

    av_col_widths = [24, 16, 28, 14, 14]
    for i, w in enumerate(av_col_widths, 1):
        ws_av.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Shifts
    # ══════════════════════════════════════════════════════════════════════════
    ws_shifts = wb.create_sheet("Shifts")

    shift_headers = ["Date", "Start Time", "End Time", "Required Skills", "Min Staff"]
    shift_hints = [
        "YYYY-MM-DD format",
        "HH:MM (24-hour)",
        "HH:MM (24-hour)",
        "Comma-separated skills required (leave blank for any employee)",
        "How many staff needed for this shift (default 1)",
    ]
    ws_shifts.append(shift_headers)
    _style_header(ws_shifts, 1, shift_headers)
    _hint_row(ws_shifts, 2, shift_hints)
    _freeze(ws_shifts, "A3")

    # Generate 7 days of sample shifts
    for i in range(7):
        d = (today + timedelta(days=i)).strftime("%Y-%m-%d")
        rows = [
            [d, "08:00", "16:00", "cashier, first_aid", 1],
            [d, "08:00", "16:00", "cashier",             2],
            [d, "14:00", "22:00", "cashier",             2],
        ]
        for row in rows:
            ws_shifts.append(row)
            r = ws_shifts.max_row
            if r % 2 == 1:
                for c in range(1, len(shift_headers) + 1):
                    ws_shifts.cell(row=r, column=c).fill = SECTION_FILL

    # Centre time / date cells
    for row in ws_shifts.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    sh_col_widths = [16, 14, 14, 38, 12]
    for i, w in enumerate(sh_col_widths, 1):
        ws_shifts.column_dimensions[get_column_letter(i)].width = w

    out = Path(__file__).parent / "frontend" / "public" / "schedule_template.xlsx"
    wb.save(out)
    print(f"Template created: {out}")


if __name__ == "__main__":
    create_template()
